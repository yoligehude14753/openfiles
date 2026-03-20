from pathlib import Path
from typing import Dict, Any
from .base import BaseParser


class XlsxParser(BaseParser):
    SUPPORTED_TYPES = ["xlsx", "xls"]

    def can_parse(self, file_type: str) -> bool:
        return file_type in self.SUPPORTED_TYPES

    def parse(self, file_path: Path) -> Dict[str, Any]:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets_data = []
            all_text = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(max_row=500, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append(cells)

                if rows:
                    header = " | ".join(rows[0]) if rows else ""
                    sheet_text = f"Sheet: {sheet_name}\n"
                    sheet_text += f"Columns: {header}\n"
                    for row in rows[1:50]:
                        sheet_text += " | ".join(row) + "\n"

                    sheets_data.append({"name": sheet_name, "rows": len(rows)})
                    all_text.append(sheet_text)

            wb.close()

            content = "\n\n".join(all_text)
            return {
                "content": content,
                "sheets": sheets_data,
                "num_sheets": len(sheets_data),
                "success": True,
            }
        except Exception as e:
            return {"content": "", "error": str(e), "success": False}
